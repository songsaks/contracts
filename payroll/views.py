from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q, Count, Sum
from decimal import Decimal, InvalidOperation
import openpyxl
from io import BytesIO

from django.contrib.auth import get_user_model, logout
from .models import WorkReport, EmployeeSalaryConfig, PayrollRecord, PayrollStatus, SSOBracket
from .forms import WorkReportForm, AdminWorkReportForm, EmployeeSalaryConfigForm

User = get_user_model()
PAYROLL_LOGIN_URL = '/payroll/login/'

def payroll_logout(request):
    """Custom logout that allows GET requests (fixes Django 5.0+ 405 error)."""
    logout(request)
    return redirect('payroll:login')

@login_required(login_url=PAYROLL_LOGIN_URL)
def login_success(request):
    """Dispatcher: redirects user to the correct dashboard based on role."""
    if request.user.is_staff:
        return redirect('payroll:admin_dashboard')
    return redirect('payroll:report_list')

# ─────────────────────────────────────────────────────────────
#  Permission Helpers — 3-Tier Role System
#  Level 1 Executive  : is_superuser
#  Level 2 HR/Admin   : is_staff (not superuser)
#  Level 3 Employee   : regular user
# ─────────────────────────────────────────────────────────────
def is_executive(user):
    """Level 1: ผู้บริหาร — superuser only."""
    return user.is_authenticated and user.is_superuser

def is_hr_or_exec(user):
    """Level 1 + 2: any staff (HR/Admin or Executive)."""
    return user.is_authenticated and user.is_staff

# Keep old name as alias so any existing references still work
is_hr = is_hr_or_exec

def payroll_members():
    """Return QS of Users who are active payroll members.
    Superusers (Executives) are ALWAYS included — no need to flag them.
    Other users must have salary_config.is_payroll_member=True."""
    from django.db.models import Q
    return User.objects.filter(
        Q(is_superuser=True) | Q(salary_config__is_payroll_member=True)
    ).distinct().filter(is_active=True).order_by('last_name', 'first_name', 'username')

def _safe_dec(value, default='0'):
    try:
        clean = str(value or default).replace(',', '').strip()
        return Decimal(clean or default)
    except (InvalidOperation, TypeError):
        return Decimal(default)

def _month_name(m):
    names = {1:'มกราคม',2:'กุมภาพันธ์',3:'มีนาคม',4:'เมษายน',5:'พฤษภาคม',6:'มิถุนายน',
             7:'กรกฎาคม',8:'สิงหาคม',9:'กันยายน',10:'ตุลาคม',11:'พฤศจิกายน',12:'ธันวาคม'}
    return names.get(m, str(m))

# ─────────────────────────────────────────────────────────────
#  Employee Views
# ─────────────────────────────────────────────────────────────
@login_required(login_url=PAYROLL_LOGIN_URL)
def report_list(request):
    reports = WorkReport.objects.filter(user=request.user).order_by('-year', '-month')
    return render(request, 'payroll/report_list.html', {'reports': reports})

@login_required(login_url=PAYROLL_LOGIN_URL)
def report_create(request):
    if request.method == 'POST':
        form = WorkReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.user = request.user
            action = request.POST.get('action', 'save')
            report.status = PayrollStatus.SUBMITTED if action == 'submit' else PayrollStatus.DRAFT
            report.save()
            if action == 'submit':
                messages.success(request, "บันทึกและส่งรายงานให้ HR เรียบร้อยแล้ว")
            else:
                messages.success(request, "บันทึกรายงานเป็นแบบร่างเรียบร้อยแล้ว")
            return redirect('payroll:report_list')
    else:
        form = WorkReportForm()
    return render(request, 'payroll/report_form.html', {'form': form, 'title': 'สร้างรายงานผลงาน', 'is_new': True})

@login_required(login_url=PAYROLL_LOGIN_URL)
def report_edit(request, pk):
    report = get_object_or_404(WorkReport, pk=pk, user=request.user)
    if report.status not in [PayrollStatus.DRAFT, PayrollStatus.REJECTED]:
        messages.error(request, "ไม่สามารถแก้ไขรายงานที่ถูกส่งหรืออนุมัติแล้วได้")
        return redirect('payroll:report_detail', pk=pk)
    if request.method == 'POST':
        form = WorkReportForm(request.POST, instance=report)
        if form.is_valid():
            r = form.save(commit=False)
            action = request.POST.get('action', 'save')
            r.status = PayrollStatus.SUBMITTED if action == 'submit' else PayrollStatus.DRAFT
            r.save()
            if action == 'submit':
                messages.success(request, "บันทึกและส่งรายงานให้ HR เรียบร้อยแล้ว")
            else:
                messages.success(request, "อัปเดตรายงานเรียบร้อยแล้ว")
            return redirect('payroll:report_list')
    else:
        form = WorkReportForm(instance=report)
    return render(request, 'payroll/report_form.html', {
        'form': form, 'report': report, 'title': 'แก้ไขรายงาน'
    })

@login_required(login_url=PAYROLL_LOGIN_URL)
def report_detail(request, pk):
    report = get_object_or_404(WorkReport, pk=pk)
    if report.user != request.user and not request.user.is_staff:
        messages.error(request, "ไม่มีสิทธิ์เข้าถึงข้อมูลนี้")
        return redirect('payroll:report_list')
    return render(request, 'payroll/report_detail.html', {'report': report})

@login_required(login_url=PAYROLL_LOGIN_URL)
def report_submit(request, pk):
    report = get_object_or_404(WorkReport, pk=pk, user=request.user)
    if report.status in [PayrollStatus.DRAFT, PayrollStatus.REJECTED]:
        report.status = PayrollStatus.SUBMITTED
        report.save()
        messages.success(request, "ส่งรายงานให้ HR ตรวจสอบเรียบร้อยแล้ว")
    else:
        messages.error(request, "ไม่สามารถส่งรายงานสถานะนี้ได้")
    return redirect('payroll:report_list')

@login_required(login_url=PAYROLL_LOGIN_URL)
def payslip_view(request, pk):
    report = get_object_or_404(WorkReport, pk=pk)
    # Employees can only see their own payslip; HR can see anyone's
    if report.user != request.user and not request.user.is_staff:
        messages.error(request, "ไม่มีสิทธิ์ดูสลิปของผู้อื่น")
        return redirect('payroll:report_list')
    # Payslip is only available after HR approves
    if report.status != PayrollStatus.APPROVED:
        messages.warning(request, "สลิปเงินเดือนนี้ยังไม่พร้อม กรุณารอ HR อนุมัติก่อน")
        return redirect('payroll:report_detail', pk=pk)
    try:
        record = report.payroll_record
    except PayrollRecord.DoesNotExist:
        messages.error(request, "ยังไม่มีผลการคำนวณเงินเดือนนี้")
        return redirect('payroll:report_detail', pk=pk)
    return render(request, 'payroll/payslip.html', {'report': report, 'record': record})

@login_required(login_url=PAYROLL_LOGIN_URL)
def my_payslips(request):
    """Employee: list all their approved payslip months."""
    approved_reports = WorkReport.objects.filter(
        user=request.user,
        status=PayrollStatus.APPROVED,
    ).select_related('payroll_record').order_by('-year', '-month')
    return render(request, 'payroll/my_payslips.html', {'reports': approved_reports})

@login_required(login_url=PAYROLL_LOGIN_URL)
def record_detail(request, pk):
    record = get_object_or_404(PayrollRecord, pk=pk)
    if record.report.user != request.user and not request.user.is_staff:
        messages.error(request, "ไม่มีสิทธิ์")
        return redirect('payroll:report_list')
    return render(request, 'payroll/payslip.html', {'report': record.report, 'record': record})

# ─────────────────────────────────────────────────────────────
#  HR / Admin Views
# ─────────────────────────────────────────────────────────────
@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def admin_dashboard(request):
    month = int(request.GET.get('month', timezone.now().month))
    year  = int(request.GET.get('year',  timezone.now().year))

    all_reports = WorkReport.objects.filter(month=month, year=year).select_related('user', 'payroll_record')
    pending     = all_reports.filter(status=PayrollStatus.SUBMITTED)
    approved    = all_reports.filter(status=PayrollStatus.APPROVED)
    total_users = payroll_members().count()  # Only payroll members, not all Django users

    context = {
        'month': month, 'year': year,
        'month_name': _month_name(month),
        'months': range(1, 13), 'years': range(2024, 2032),
        'all_reports': all_reports,
        'pending': pending,
        'approved': approved,
        'total_users': total_users,
        'submitted_count': pending.count(),
        'approved_count': approved.count(),
        'is_exec': request.user.is_superuser,
    }
    return render(request, 'payroll/admin_dashboard.html', context)

@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def admin_approve(request, pk):
    report = get_object_or_404(WorkReport, pk=pk)
    config, _ = EmployeeSalaryConfig.objects.get_or_create(user=report.user)

    if request.method == 'POST':
        # Status comes from the button's "name=status value=APPROVED/REJECTED"
        new_status = request.POST.get('status', '').strip().upper()
        admin_remarks = request.POST.get('admin_remarks', '').strip()

        if new_status in (PayrollStatus.APPROVED, PayrollStatus.REJECTED):
            report.status = new_status
            report.admin_remarks = admin_remarks
            report.save()

            if new_status == PayrollStatus.APPROVED:
                _calculate_and_save_payroll(report, config, request.user)
                messages.success(
                    request,
                    f"✅ อนุมัติและคำนวณเงินเดือน {report.user.get_full_name() or report.user.username} เรียบร้อยแล้ว"
                )
            else:
                messages.warning(
                    request,
                    f"↩️ ส่งกลับรายงานของ {report.user.get_full_name() or report.user.username} ให้แก้ไข"
                )
            return redirect('payroll:admin_dashboard')
        else:
            messages.error(request, "กรุณาเลือกสถานะ: อนุมัติ หรือ ส่งกลับแก้ไข")

    form = AdminWorkReportForm(instance=report)

    review_data = [
        ("วันทำงาน", f"{report.working_days} วัน"),
        ("OT", f"{report.ot_hours} ชม."),
        ("คอมมิชชัน", f"{report.commissions:,.0f} ฿"),
        ("Incentives", f"{report.incentives:,.0f} ฿"),
        ("PB Score", f"{report.pb_liva_score}"),
        ("ค่าบริหารทีม", f"{report.team_mgmt_fee:,.0f} ฿"),
        ("ค่าวิชาชีพ", f"{report.professional_fee:,.0f} ฿"),
        ("วันขาดงาน", f"{report.absent_days} วัน"),
        ("หักขาดงาน", f"{report.absent_deduction_amount:,.0f} ฿"),
        ("เบิกล่วงหน้า", f"{report.advance_pay:,.0f} ฿"),
        ("ออมทรัพย์", f"{report.savings:,.0f} ฿"),
        ("หมายเหตุพนักงาน", report.remarks or "—"),
    ]
    return render(request, 'payroll/admin_approve.html', {
        'form': form, 'report': report, 'config': config,
        'review_data': review_data,
    })

def _calculate_and_save_payroll(report, config, processed_by):
    """Central payroll calculation logic."""
    base    = config.base_salary
    ot_pay  = report.ot_hours * config.ot_rate_per_hour
    total_income = (base + ot_pay + report.team_mgmt_fee + report.professional_fee
                    + report.commissions + report.incentives)

    ss_amt = config.get_sso_amount()
    total_deductions = (ss_amt + config.tax_withholding + report.absent_deduction_amount
                        + report.advance_pay + report.savings + report.lost_equipment_fee
                        + report.other_deductions)

    PayrollRecord.objects.update_or_create(
        report=report,
        defaults={
            'base_salary_snapshot': base,
            'ot_amount': ot_pay,
            'social_security_amount': ss_amt,
            'tax_amount': config.tax_withholding,
            'total_income': total_income,
            'total_deductions': total_deductions,
            'net_pay': total_income - total_deductions,
            'processed_by': processed_by,
        }
    )

def _preview_payroll(report, config):

    """Calculate payroll figures WITHOUT saving — used for batch preview."""
    base         = config.base_salary
    ot_pay       = report.ot_hours * config.ot_rate_per_hour
    total_income = (base + ot_pay + report.team_mgmt_fee + report.professional_fee
                    + report.commissions + report.incentives)
    ss_amt       = config.get_sso_amount()
    total_ded    = (ss_amt + config.tax_withholding + report.absent_deduction_amount
                    + report.advance_pay + report.savings
                    + report.lost_equipment_fee + report.other_deductions)
    return {
        'base': base, 'ot_pay': ot_pay, 'total_income': total_income,
        'ss': ss_amt, 'tax': config.tax_withholding,
        'total_ded': total_ded, 'net_pay': total_income - total_ded,
        'bank_account': config.bank_account_number,
        'bank_name': config.bank_name,
    }

# ── Bank Transfer Export (Executive Only) ─────────────────
@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def bank_export(request):
    """Screen: review approved payroll for a month, mark paid, download bank file."""
    from datetime import date as date_cls
    month = int(request.GET.get('month', timezone.now().month))
    year  = int(request.GET.get('year',  timezone.now().year))

    if request.method == 'POST':
        action     = request.POST.get('action', '')
        pay_date   = request.POST.get('payment_date') or str(date_cls.today())
        record_ids = request.POST.getlist('record_ids')

        if action == 'mark_paid' and record_ids:
            PayrollRecord.objects.filter(pk__in=record_ids).update(
                is_paid=True,
                payment_date=pay_date,
            )
            messages.success(request, f"✅ บันทึกสถานะจ่ายแล้ว {len(record_ids)} รายการ (วันที่ {pay_date})")
        return redirect(f"{request.path}?month={month}&year={year}")

    records = (PayrollRecord.objects
               .filter(report__month=month, report__year=year)
               .select_related('report__user', 'report__user__salary_config')
               .order_by('report__user__last_name', 'report__user__first_name'))

    rows = []
    total_net = Decimal('0')
    for rec in records:
        cfg = getattr(rec.report.user, 'salary_config', None)
        rows.append({
            'record': rec,
            'user': rec.report.user,
            'bank_name': cfg.bank_name if cfg else '—',
            'bank_account': cfg.bank_account_number if cfg else '—',
            'net_pay': rec.net_pay,
            'is_paid': rec.is_paid,
            'payment_date': rec.payment_date,
        })
        total_net += rec.net_pay

    context = {
        'month': month, 'year': year,
        'month_name': _month_name(month),
        'months': range(1, 13), 'years': range(2024, 2032),
        'rows': rows,
        'total_net': total_net,
        'count': len(rows),
        'today': date_cls.today().isoformat(),
    }
    return render(request, 'payroll/bank_export.html', context)


@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def bank_export_excel(request):
    """Download approved payroll as Excel for bank upload (KBank / generic)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    month  = int(request.GET.get('month', timezone.now().month))
    year   = int(request.GET.get('year',  timezone.now().year))
    fmt    = request.GET.get('format', 'generic')   # 'generic' | 'kbank'

    records = (PayrollRecord.objects
               .filter(report__month=month, report__year=year)
               .select_related('report__user', 'report__user__salary_config')
               .order_by('report__user__last_name', 'report__user__first_name'))

    wb = openpyxl.Workbook()
    ws = wb.active

    green = PatternFill("solid", fgColor="064E3B")
    white_bold = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alt_fill = PatternFill("solid", fgColor="F0FDF4")
    total_fill = PatternFill("solid", fgColor="065F46")

    if fmt == 'kbank':
        ws.title = f"KBank_{month:02d}_{year}"
        # KBank Business Online format
        # SEQ | DEBIT_REF | CREDIT_BANK | CREDIT_ACC | CREDIT_NAME | AMOUNT | EMAIL | SMS | NOTE
        headers = ["SEQ","DEBIT_REF","CREDIT_BANK_CODE","CREDIT_ACC","CREDIT_NAME","AMOUNT","EMAIL","SMS","NOTE"]
        col_widths = [6, 15, 15, 20, 30, 15, 25, 15, 20]
    else:
        ws.title = f"Payroll_{month:02d}_{year}"
        headers = [
            "ลำดับ","ชื่อ-นามสกุล","ธนาคาร","เลขที่บัญชี",
            "ยอดโอนสุทธิ (บาท)","หมายเหตุ","สถานะ"
        ]
        col_widths = [8, 30, 20, 20, 18, 25, 12]

    # Write header
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = white_bold
        cell.fill = green
        cell.alignment = header_align
        cell.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    total_net = Decimal('0')
    for seq, rec in enumerate(records, 1):
        cfg  = getattr(rec.report.user, 'salary_config', None)
        name = rec.report.user.get_full_name() or rec.report.user.username
        bank_name = cfg.bank_name if cfg else ''
        bank_acc  = cfg.bank_account_number if cfg else ''
        amount    = float(rec.net_pay)
        total_net += rec.net_pay

        if fmt == 'kbank':
            row_data = [
                seq,
                f"PAY{year}{month:02d}",
                '',                 # CREDIT_BANK_CODE: fill per bank (e.g. 004=KBank)
                bank_acc,
                name,
                amount,
                rec.report.user.email or '',
                '',
                f"เงินเดือน {_month_name(month)} {year}",
            ]
        else:
            row_data = [
                seq, name, bank_name, bank_acc,
                amount,
                f"เงินเดือน {_month_name(month)} {year}",
                "จ่ายแล้ว" if rec.is_paid else "รอโอน",
            ]

        fill = alt_fill if seq % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=seq + 1, column=ci, value=val)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            # Format amount column
            if fmt == 'kbank' and ci == 6:
                cell.number_format = '#,##0.00'
            elif fmt == 'generic' and ci == 5:
                cell.number_format = '#,##0.00'

    # Total row
    total_row = len(list(records)) + 2
    ws.cell(row=total_row, column=1 if fmt == 'kbank' else 1, value="รวม").font = Font(bold=True, color="FFFFFF")

    if fmt == 'kbank':
        tc = ws.cell(row=total_row, column=6, value=float(total_net))
        tc.number_format = '#,##0.00'
    else:
        tc = ws.cell(row=total_row, column=5, value=float(total_net))
        tc.number_format = '#,##0.00'
        tc.font = Font(bold=True, color="FFFFFF")

    for ci in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=ci).fill = total_fill
        ws.cell(row=total_row, column=ci).border = thin

    filename = f"payroll_bank_{fmt}_{year}{month:02d}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


# ── Batch Approve (all employees at once) ──────────────────
@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def batch_approve(request):
    month = int(request.GET.get('month', timezone.now().month))
    year  = int(request.GET.get('year',  timezone.now().year))

    if request.method == 'POST':
        action   = request.POST.get('action', '')        # 'approve_all' | 'approve_one' | 'reject_one'
        report_ids = request.POST.getlist('report_ids')  # for approve_all
        single_id  = request.POST.get('report_id')       # for single actions

        approved = rejected = 0

        if action == 'approve_all':
            for rid in report_ids:
                try:
                    r = WorkReport.objects.get(pk=rid, status=PayrollStatus.SUBMITTED)
                    cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=r.user)
                    r.status = PayrollStatus.APPROVED
                    r.save()
                    _calculate_and_save_payroll(r, cfg, request.user)
                    approved += 1
                except WorkReport.DoesNotExist:
                    pass
            messages.success(request, f"✅ อนุมัติพนักงาน {approved} คน คำนวณเงินเดือนเรียบร้อยแล้ว")

        elif action == 'approve_one' and single_id:
            try:
                r = WorkReport.objects.get(pk=single_id)
                cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=r.user)
                r.status = PayrollStatus.APPROVED
                r.admin_remarks = request.POST.get('admin_remarks', '')
                r.save()
                _calculate_and_save_payroll(r, cfg, request.user)
                messages.success(request, f"✅ อนุมัติ {r.user.get_full_name() or r.user.username} แล้ว")
            except WorkReport.DoesNotExist:
                messages.error(request, "ไม่พบรายงาน")

        elif action == 'reject_one' and single_id:
            try:
                r = WorkReport.objects.get(pk=single_id)
                r.status = PayrollStatus.REJECTED
                r.admin_remarks = request.POST.get('admin_remarks', '')
                r.save()
                messages.warning(request, f"↩️ ส่งกลับ {r.user.get_full_name() or r.user.username} แล้ว")
            except WorkReport.DoesNotExist:
                messages.error(request, "ไม่พบรายงาน")

        return redirect(f"{request.path}?month={month}&year={year}")

    # GET: build preview rows
    submitted = (WorkReport.objects
                 .filter(month=month, year=year, status=PayrollStatus.SUBMITTED)
                 .select_related('user')
                 .order_by('user__last_name', 'user__first_name'))
    approved  = (WorkReport.objects
                 .filter(month=month, year=year, status=PayrollStatus.APPROVED)
                 .select_related('user', 'payroll_record')
                 .order_by('user__last_name', 'user__first_name'))

    pending_rows = []
    for r in submitted:
        cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=r.user)
        pending_rows.append({'report': r, 'cfg': cfg, 'preview': _preview_payroll(r, cfg)})

    approved_rows = []
    total_net = Decimal('0')
    for r in approved:
        try:
            rec = r.payroll_record
            approved_rows.append({'report': r, 'record': rec})
            total_net += rec.net_pay
        except PayrollRecord.DoesNotExist:
            pass

    # Missing rows: payroll members who have NO report for this month (exec can create for them)
    reports_this_month_user_ids = set(
        WorkReport.objects.filter(month=month, year=year).values_list('user_id', flat=True)
    )
    missing_users = payroll_members().filter(is_active=True).exclude(id__in=reports_this_month_user_ids)
    missing_rows = []
    for u in missing_users:
        cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=u)
        missing_rows.append({'user': u, 'cfg': cfg})

    context = {
        'month': month, 'year': year,
        'month_name': _month_name(month),
        'months': range(1, 13), 'years': range(2024, 2032),
        'pending_rows': pending_rows,
        'approved_rows': approved_rows,
        'missing_rows': missing_rows,
        'pending_count': len(pending_rows),
        'approved_count': len(approved_rows),
        'missing_count': len(missing_rows),
        'total_net': total_net,
        'grand_total_income': sum(r['record'].total_income for r in approved_rows),
        'grand_total_ded': sum(r['record'].total_deductions for r in approved_rows),
    }
    return render(request, 'payroll/batch_approve.html', context)


# ── Exec: Create Report for a missing employee inline ─────
@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def exec_create_report(request):
    """Executive creates/updates a WorkReport for any employee on the batch_approve page."""
    if request.method != 'POST':
        return redirect('payroll:batch_approve')
    month   = int(request.POST.get('month', timezone.now().month))
    year    = int(request.POST.get('year',  timezone.now().year))
    user_id = request.POST.get('user_id')
    user    = get_object_or_404(User, id=user_id)

    def dec(k): return _safe_dec(request.POST.get(k, '0'))

    report, _ = WorkReport.objects.get_or_create(
        user=user, month=month, year=year,
        defaults={'status': PayrollStatus.SUBMITTED}
    )
    report.working_days            = dec('working_days')
    report.ot_hours                = dec('ot_hours')
    report.commissions             = dec('commissions')
    report.incentives              = dec('incentives')
    report.pb_liva_score           = dec('pb_liva_score')
    report.team_mgmt_fee           = dec('team_mgmt_fee')
    report.professional_fee        = dec('professional_fee')
    report.absent_days             = dec('absent_days')
    report.absent_deduction_amount = dec('absent_deduction')
    report.advance_pay             = dec('advance_pay')
    report.savings                 = dec('savings')
    report.status = PayrollStatus.SUBMITTED
    report.save()

    # If exec chose to approve immediately
    if request.POST.get('and_approve') == '1':
        cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=user)
        report.status = PayrollStatus.APPROVED
        report.save()
        _calculate_and_save_payroll(report, cfg, request.user)
        messages.success(request, f"✅ บันทึกและอนุมัติ {user.get_full_name() or user.username} แล้ว")
    else:
        messages.success(request, f"✅ บันทึกข้อมูล {user.get_full_name() or user.username} เรียบร้อยแล้ว")
    return redirect(f"{reverse('payroll:batch_approve')}?month={month}&year={year}")


# ── Salary Config (Executive Only) ────────────────────────
@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def salary_config_list(request):
    # Only payroll members
    members = payroll_members()
    for u in members:
        EmployeeSalaryConfig.objects.get_or_create(user=u)
    configs = EmployeeSalaryConfig.objects.filter(
        is_payroll_member=True
    ).select_related('user').order_by('user__last_name', 'user__first_name')
    return render(request, 'payroll/salary_config_list.html', {'configs': configs})

@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def salary_config_edit(request, user_id):
    user_obj = get_object_or_404(User, id=user_id)
    config, _ = EmployeeSalaryConfig.objects.get_or_create(user=user_obj)
    if request.method == 'POST':
        form = EmployeeSalaryConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, f"อัปเดตโครงสร้างเงินเดือนของ {user_obj.get_full_name() or user_obj.username} เรียบร้อยแล้ว")
            return redirect('payroll:salary_config_list')
    else:
        form = EmployeeSalaryConfigForm(instance=config)
    return render(request, 'payroll/salary_config_form.html', {'form': form, 'user_obj': user_obj})


# ── SSO Bracket Management (HR + Exec) ────────────────────
@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def sso_bracket_config(request):
    """HR/Admin and Exec can manage progressive SSO rate brackets."""
    if request.method == 'POST':
        action = request.POST.get('action', '')
        if action == 'add':
            try:
                SSOBracket.objects.create(
                    min_salary  = Decimal(request.POST.get('min_salary', '0') or '0'),
                    max_salary  = Decimal(request.POST.get('max_salary')) if request.POST.get('max_salary') else None,
                    rate_percent = Decimal(request.POST.get('rate_percent', '5')),
                    salary_cap  = Decimal(request.POST.get('salary_cap', '0') or '0'),
                    description = request.POST.get('description', '').strip(),
                )
                messages.success(request, "✅ เพิ่มขั้นบันไดประกันสังคมเรียบร้อยแล้ว")
            except Exception as e:
                messages.error(request, f"เกิดข้อผิดพลาด: {e}")
        elif action == 'toggle':
            bracket = get_object_or_404(SSOBracket, pk=request.POST.get('bracket_id'))
            bracket.is_active = not bracket.is_active
            bracket.save()
            messages.success(request, f"{'เปิด' if bracket.is_active else 'ปิด'}ใช้งาน bracket เรียบร้อย")
        elif action == 'delete':
            SSOBracket.objects.filter(pk=request.POST.get('bracket_id')).delete()
            messages.success(request, "ลบ bracket เรียบร้อยแล้ว")
        return redirect('payroll:sso_bracket_config')

    brackets = SSOBracket.objects.all()
    # Preview SSO for each payroll member using current brackets
    previews = []
    for cfg in EmployeeSalaryConfig.objects.filter(is_payroll_member=True).select_related('user'):
        previews.append({
            'name': cfg.user.get_full_name() or cfg.user.username,
            'code': cfg.employee_code or '—',
            'base': cfg.base_salary,
            'sso_amt': cfg.get_sso_amount(),
        })
    # Also include superusers
    from django.db.models import Q as _Q
    for cfg in EmployeeSalaryConfig.objects.filter(user__is_superuser=True).select_related('user'):
        if not any(p['name'] == (cfg.user.get_full_name() or cfg.user.username) for p in previews):
            previews.append({
                'name': cfg.user.get_full_name() or cfg.user.username,
                'code': cfg.employee_code or '—',
                'base': cfg.base_salary,
                'sso_amt': cfg.get_sso_amount(),
            })
    previews.sort(key=lambda x: x['name'])
    return render(request, 'payroll/sso_bracket_config.html', {
        'brackets': brackets,
        'previews': previews,
        'is_exec': request.user.is_superuser,
    })

# ── User Management (Executive Only) ──────────────────────
@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def user_management(request):
    q = request.GET.get('q', '')
    users = payroll_members()
    if q:
        users = users.filter(
            Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q)
        )
    return render(request, 'payroll/user_management.html', {'users': users, 'q': q})

@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def toggle_staff(request, user_id):
    if request.method == 'POST':
        target = get_object_or_404(User, id=user_id)
        if target == request.user:
            messages.error(request, "ไม่สามารถเปลี่ยนสิทธิ์ตัวเองได้")
        else:
            # Cycle: employee → HR → exec → employee
            if not target.is_staff:
                target.is_staff = True
                target.is_superuser = False
                role = "HR/Admin"
            elif target.is_staff and not target.is_superuser:
                target.is_superuser = True
                role = "ผู้บริหาร (Executive)"
            else:
                target.is_staff = False
                target.is_superuser = False
                role = "พนักงาน"
            target.save()
            messages.success(request, f"เปลี่ยนสิทธิ์ {target.username} เป็น {role}")
    return redirect('payroll:user_management')

@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def set_user_password(request, user_id):
    """Executive: reset/set password for any user directly from User Management."""
    if request.method != 'POST':
        return redirect('payroll:user_management')
    target = get_object_or_404(User, id=user_id)
    if target == request.user:
        messages.error(request, "ใช้ Django Admin เพื่อเปลี่ยน password ของตัวเอง")
        return redirect('payroll:user_management')
    new_pass = request.POST.get('new_password', '').strip()
    if len(new_pass) < 6:
        messages.error(request, "รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร")
        return redirect('payroll:user_management')
    target.set_password(new_pass)
    target.save()
    messages.success(request,
        f"✅ เปลี่ยนรหัสผ่านของ {target.get_full_name() or target.username} เรียบร้อยแล้ว")
    return redirect('payroll:user_management')

# ── Excel-Grid Bulk Entry (HR & Exec) ─────────────────────
@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def bulk_management(request):
    month = int(request.GET.get('month', timezone.now().month))
    year  = int(request.GET.get('year',  timezone.now().year))

    # Exec sees everyone; HR sees only non-executive employees
    if request.user.is_superuser:
        users = payroll_members().filter(is_active=True)
    else:
        # HR/Admin cannot see executives in the data entry grid
        users = payroll_members().filter(is_active=True, is_superuser=False)
    reports_map = {r.user_id: r for r in WorkReport.objects.filter(month=month, year=year)}

    if request.method == 'POST':
        saved = 0
        for user in users:
            p = f"u{user.id}_"
            def pv(key): return request.POST.get(f"{p}{key}", "0")
            report, _ = WorkReport.objects.get_or_create(
                user=user, month=month, year=year,
                defaults={'status': PayrollStatus.SUBMITTED}
            )
            report.working_days            = _safe_dec(pv('working_days'))
            report.ot_hours                = _safe_dec(pv('ot_hours'))
            report.commissions             = _safe_dec(pv('commissions'))
            report.incentives              = _safe_dec(pv('incentives'))
            report.pb_liva_score           = _safe_dec(pv('pb_liva_score'))
            report.team_mgmt_fee           = _safe_dec(pv('team_mgmt_fee'))
            report.professional_fee        = _safe_dec(pv('professional_fee'))
            report.absent_days             = _safe_dec(pv('absent_days'))
            report.absent_deduction_amount = _safe_dec(pv('absent_deduction'))
            report.advance_pay             = _safe_dec(pv('advance_pay'))
            report.savings                 = _safe_dec(pv('savings'))
            report.save()
            saved += 1
        messages.success(request, f"บันทึกข้อมูลพนักงาน {saved} คน สำหรับเดือน {_month_name(month)} {year} เรียบร้อยแล้ว")
        return redirect(f"{request.path}?month={month}&year={year}")

    return render(request, 'payroll/bulk_management.html', {
        'users': users,
        'reports_map': reports_map,
        'month': month, 'year': year,
        'month_name': _month_name(month),
        'months': range(1, 13),
        'years': range(2024, 2032),
    })

@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def bulk_save_row(request):
    """AJAX: save a single employee row without page reload."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error'}, status=405)

    user_id = request.POST.get('user_id')
    month   = int(request.POST.get('month', 1))
    year    = int(request.POST.get('year',  timezone.now().year))

    user = get_object_or_404(User, id=user_id)
    report, _ = WorkReport.objects.get_or_create(
        user=user, month=month, year=year,
        defaults={'status': PayrollStatus.SUBMITTED}
    )

    def dec(k): return _safe_dec(request.POST.get(k))

    report.working_days            = dec('working_days')
    report.ot_hours                = dec('ot_hours')
    report.commissions             = dec('commissions')
    report.incentives              = dec('incentives')
    report.pb_liva_score           = dec('pb_liva_score')
    report.team_mgmt_fee           = dec('team_mgmt_fee')
    report.professional_fee        = dec('professional_fee')
    report.absent_days             = dec('absent_days')
    report.absent_deduction_amount = dec('absent_deduction')
    report.advance_pay             = dec('advance_pay')
    report.savings                 = dec('savings')
    report.save()

    return JsonResponse({
        'status': 'success',
        'updated_at': report.updated_at.strftime("%H:%M"),
        'report_status': report.get_status_display(),
    })

# ── Excel Template Download (HR & Exec) ───────────────────
@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def download_template(request):
    month = int(request.GET.get('month', timezone.now().month))
    year  = int(request.GET.get('year',  timezone.now().year))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll_{month}_{year}"

    headers = [
        "Username","ชื่อ","นามสกุล","Month","Year",
        "WorkingDays","OT_Hours","Commissions","Incentives","PBLivaScore",
        "TeamMgmtFee","ProfessionalFee",
        "AbsentDays","AbsentDeduction","AdvancePay","Savings"
    ]
    ws.append(headers)
    for user in User.objects.filter(is_active=True).order_by('last_name','first_name'):
        ws.append([user.username, user.first_name, user.last_name, month, year,
                   0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=payroll_{month}_{year}.xlsx'
    wb.save(response)
    return response

# ── Excel Import (HR & Exec) ──────────────────────────────
@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def import_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            ws = wb.active
            ok, skip = 0, 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                username = str(row[0]).strip()
                try:
                    user   = User.objects.get(username=username)
                    month  = int(row[3] or timezone.now().month)
                    year   = int(row[4] or timezone.now().year)
                    r, _   = WorkReport.objects.get_or_create(
                        user=user, month=month, year=year,
                        defaults={'status': PayrollStatus.SUBMITTED}
                    )
                    r.working_days            = _safe_dec(row[5])
                    r.ot_hours                = _safe_dec(row[6])
                    r.commissions             = _safe_dec(row[7])
                    r.incentives              = _safe_dec(row[8])
                    r.pb_liva_score           = _safe_dec(row[9])
                    r.team_mgmt_fee           = _safe_dec(row[10])
                    r.professional_fee        = _safe_dec(row[11])
                    r.absent_days             = _safe_dec(row[12])
                    r.absent_deduction_amount = _safe_dec(row[13])
                    r.advance_pay             = _safe_dec(row[14])
                    r.savings                 = _safe_dec(row[15])
                    r.status = PayrollStatus.SUBMITTED
                    r.save()
                    ok += 1
                except User.DoesNotExist:
                    skip += 1
            messages.success(request, f"นำเข้าสำเร็จ {ok} รายการ (ข้าม {skip} รายการที่หา user ไม่เจอ)")
        except Exception as e:
            messages.error(request, f"เกิดข้อผิดพลาด: {e}")
    return redirect('payroll:bulk_management')

# ── Employee List & Add Single Employee (HR & Exec) ───────
@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def employee_list(request):
    """Show payroll members only, with add-single and import tools."""
    q = request.GET.get('q', '')
    qs = payroll_members()
    if q:
        qs = qs.filter(
            Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q)
        )
    # Prefetch salary config (for bank info)
    qs = qs.prefetch_related('salary_config')
    return render(request, 'payroll/employee_list.html', {
        'employees': qs,
        'q': q,
        'total': qs.count(),
        'is_exec': request.user.is_superuser,
    })

@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def create_payroll_employee(request):
    """HR or Exec: add a single new employee to the payroll system."""
    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '').strip() or '12345678'
        make_hr    = request.POST.get('make_hr') == '1'

        if not username:
            messages.error(request, "กรุณาใส่ Username")
            return redirect('payroll:employee_list')
        if ' ' in username:
            messages.error(request, "Username ต้องไม่มีช่องว่าง")
            return redirect('payroll:employee_list')
        if User.objects.filter(username=username).exists():
            u = User.objects.get(username=username)
        else:
            u = User.objects.create_user(
                username=username, password=password,
                first_name=first_name, last_name=last_name, email=email,
            )
        cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=u)
        cfg.is_payroll_member = True
        cfg.bank_name           = request.POST.get('bank_name', '').strip()
        cfg.bank_account_number = request.POST.get('bank_account_number', '').strip()
        # Auto-generate employee_code if not set
        if not cfg.employee_code:
            cfg.employee_code = EmployeeSalaryConfig.generate_employee_code()
        cfg.save()
        if make_hr and request.user.is_superuser:
            u.is_staff = True
            u.save()
        messages.success(request,
            f"✅ เพิ่ม {u.get_full_name() or u.username} เข้าระบบ Payroll เรียบร้อยแล้ว")
    return redirect('payroll:employee_list')

@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def edit_payroll_employee(request, user_id):
    """HR or Exec: edit basic info of a payroll member."""
    target = get_object_or_404(
        User, id=user_id, salary_config__is_payroll_member=True
    )
    if request.method == 'POST':
        target.first_name = request.POST.get('first_name', '').strip()
        target.last_name  = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        if email:
            target.email = email
        active_val = request.POST.get('is_active', '1')
        target.is_active = active_val == '1'
        target.save()
        # Update bank info in salary config
        cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=target)
        cfg.bank_name           = request.POST.get('bank_name', '').strip()
        cfg.bank_account_number = request.POST.get('bank_account_number', '').strip()
        cfg.save()
        messages.success(request,
            f"✅ เปลี่ยนข้อมูล {target.get_full_name() or target.username} เรียบร้อยแล้ว")
    return redirect('payroll:employee_list')

@user_passes_test(is_executive, login_url=PAYROLL_LOGIN_URL)
def remove_payroll_member(request, user_id):
    """Exec only: remove a user from payroll (does not delete the Django user)."""
    if request.method == 'POST':
        target = get_object_or_404(User, id=user_id)
        cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=target)
        cfg.is_payroll_member = False
        cfg.save()
        messages.warning(request,
            f"⚠️ นำ {target.get_full_name() or target.username} ออกจากระบบ Payroll แล้ว")
    return redirect('payroll:employee_list')

@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def download_employee_template(request):
    """Generate an Excel template for importing employee roster."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายชื่อพนักงาน"

    # Header row styling
    headers = [
        ("username*",             "ชื่อผู้ใช้ (ห้ามซ้ำ, ไม่มีช่องว่าง)"),  # A
        ("first_name*",           "ชื่อจริง"),                               # B
        ("last_name*",            "นามสกุล"),                                # C
        ("email",                 "อีเมล (ไม่บังคับ)"),                      # D
        ("password",              "รหัสผ่านเริ่มต้น (เว้นว่าง = '12345678')"), # E
        ("bank_name",             "ธนาคาร เช่น KBank / กสิกร"),             # F
        ("bank_account_number",   "เลขบัญชี เช่น 012-3-45678-9"),           # G
        ("is_active",             "ใช้งาน? (TRUE/FALSE)"),                   # H
    ]

    # Style header
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = [20, 20, 20, 30, 25, 25, 12, 20]
    for i, (col_key, col_label) in enumerate(headers, 1):
        cell_key   = ws.cell(row=1, column=i, value=col_key)
        cell_label = ws.cell(row=2, column=i, value=col_label)
        for cell in (cell_key, cell_label):
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    # Example rows
    examples = [
        ("emp001", "สมชาย", "ใจดี", "somchai@company.com", "", "", "TRUE", "ตัวอย่าง"),
        ("emp002", "สมหญิง", "รักดี", "", "mypass123", "", "TRUE", ""),
    ]
    example_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    for r_idx, ex in enumerate(examples, 3):
        for c_idx, val in enumerate(ex, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = example_fill

    # Pre-fill existing employees
    existing_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    for r_idx, u in enumerate(payroll_members().order_by('last_name', 'first_name'), len(examples) + 3):
        cfg = getattr(u, 'salary_config', None)
        ws.cell(row=r_idx, column=1, value=u.username)
        ws.cell(row=r_idx, column=2, value=u.first_name)
        ws.cell(row=r_idx, column=3, value=u.last_name)
        ws.cell(row=r_idx, column=4, value=u.email)
        ws.cell(row=r_idx, column=6, value=cfg.bank_name if cfg else '')
        ws.cell(row=r_idx, column=7, value=cfg.bank_account_number if cfg else '')
        ws.cell(row=r_idx, column=8, value="TRUE" if u.is_active else "FALSE")
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).fill = existing_fill

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_import_template.xlsx'
    wb.save(response)
    return response

@user_passes_test(is_hr_or_exec, login_url=PAYROLL_LOGIN_URL)
def import_employees(request):
    """Import /update employee (User) records from an uploaded Excel file."""
    if request.method != 'POST' or not request.FILES.get('excel_file'):
        return redirect('payroll:employee_list')

    try:
        wb = openpyxl.load_workbook(request.FILES['excel_file'])
        ws = wb.active

        created_count = updated_count = skipped_count = 0
        errors = []

        # rows start at row 3 (rows 1-2 are header/label)
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            username = str(row[0] or '').strip()
            if not username:
                continue

            first_name    = str(row[1] or '').strip()
            last_name     = str(row[2] or '').strip()
            email         = str(row[3] or '').strip()
            password      = str(row[4] or '').strip() or '12345678'
            bank_name     = str(row[5] or '').strip()   # col F
            bank_acc_no   = str(row[6] or '').strip()   # col G
            is_active_raw = str(row[7] or 'TRUE').strip().upper()
            is_active  = is_active_raw in ('TRUE', '1', 'YES', 'Y', 'จริง', 'ใช้งาน')

            # Validate username
            if ' ' in username:
                errors.append(f"แถว {row_num}: username '{username}' มีช่องว่าง — ข้ามแถวนี้")
                skipped_count += 1
                continue

            try:
                user = User.objects.get(username=username)
                # Update existing
                user.first_name = first_name
                user.last_name  = last_name
                if email:
                    user.email = email
                user.is_active = is_active
                user.save()
                updated_count += 1
            except User.DoesNotExist:
                # Create new user
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    is_active=is_active,
                )
                # Auto-create salary config and mark as payroll member
                cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=user)
                cfg.is_payroll_member   = True
                cfg.bank_name           = bank_name
                cfg.bank_account_number = bank_acc_no
                cfg.save()
                created_count += 1
            else:
                # Existing user: also ensure they're marked as payroll member + update bank
                cfg, _ = EmployeeSalaryConfig.objects.get_or_create(user=user)
                cfg.is_payroll_member = True
                if bank_name:   cfg.bank_name           = bank_name
                if bank_acc_no: cfg.bank_account_number = bank_acc_no
                cfg.save()

        msg = f"✅ สร้างใหม่ {created_count} คน, อัปเดต {updated_count} คน"
        if skipped_count:
            msg += f", ข้าม {skipped_count} แถว"
        messages.success(request, msg)

        if errors:
            for err in errors[:5]:  # show max 5 errors
                messages.warning(request, err)

    except Exception as e:
        messages.error(request, f"เกิดข้อผิดพลาดในการอ่านไฟล์: {e}")

    return redirect('payroll:employee_list')
